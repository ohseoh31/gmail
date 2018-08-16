import os
from openpyxl import Workbook
import hashlib
import datetime

#파일 경로 검색
def search(dirname):
    fileName_list = []
    try:
        filenames = os.listdir(dirname)
        for filename in filenames:
            full_filename = os.path.join(dirname, filename)
            #print ("full : ",full_filename)
            if os.path.isdir(full_filename):
                search(full_filename)
            else:
                ext = os.path.splitext(full_filename)[-1]
                if ext == '.jpg' or ext == '.png' or ext == '.bmp':

                    fileName_list.append(full_filename)
                    #print(fileName_list)
        return fileName_list

    except PermissionError:
        print("permission Denided")
        pass

def appendExcel(outpath, info):
    wb = Workbook.load_workbook('data.xlsx')
    ws = wb.active
    max = ws.max_row

    print (max)
#CSV 엑셀파일 저장
def saveExcel(outpath, info):

    wb = Workbook()
    ws = wb.active
    ws.titles = "download Img info"
    ws['A1'] = '시간'
    ws['B1'] = '파일명'
    ws['C1'] = 'url 정보'
    ws['D1'] = '단축 url'
    ws['E1'] = 'GPS 정보'
    ws['F1'] = 'MD5'
    ws['G1'] = 'SHA1'
    column = 1
    for i in info :
        for row in range (2, len(i)+2):
            if i[row-2] == '[]':
                print(i[row-2])
                continue
            ws.cell(row=row, column= column, value=str(i[row-2]))

        column = column + 1
    now = datetime.datetime.now()
    date = str(now.year) + "_" + str(now.month) + "_" + str(now.day)
    file = date + '.xlsx'
    wb.save(outpath + file)
    #os.system("start %s" % (outpath + file))
    print ('excel_save')


#파일 MD5 해시값 얻기
def getHash_MD5(path, blocksize=65536):
    afile = open(path, 'rb')
    hasher = hashlib.md5()
    buf = afile.read(blocksize)
    while len(buf) > 0:
        hasher.update(buf)
        buf = afile.read(blocksize)
    afile.close()
    return hasher.hexdigest()

#파일 SHA1 해시값 얻기
def getHash_SHA1(path, blocksize=65536):
    afile = open(path, 'rb')
    hasher = hashlib.sha1()
    buf = afile.read(blocksize)
    while len(buf) > 0:
        hasher.update(buf)
        buf = afile.read(blocksize)
    afile.close()
    return hasher.hexdigest()