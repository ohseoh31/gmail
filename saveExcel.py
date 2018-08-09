import os
from openpyxl import Workbook


#파일 경로 검색
def search(dirname):
    fileName_list = []
    try:
        filenames = os.listdir(dirname)
        for filename in filenames:
            full_filename = os.path.join(dirname, filename)
            if os.path.isdir(full_filename):
                search(full_filename)
            else:
                ext = os.path.splitext(full_filename)[-1]
                if ext == '.jpg' or ext == '.png' or ext == '.bmp':

                    fileName_list.append(full_filename)

        return fileName_list

    except PermissionError:
        print("permission Denided")
        pass

#CSV 엑셀파일 저장
def saveExcel(outpath,url,fileName):
    wb = Workbook()
    ws = wb.active
    ws.title = "download Img info"
    ws['A1'] = 'url 정보'
    ws['B1'] = '파일명'
    for i in range(0, len(url)):
        url_text = 'A' + str(i+2)
        ws[url_text] = url[i]

    for i in range(0, len(fileName)):
        fileName_text = 'B' + str(i+2)
        ws[fileName_text] = fileName[i]

    now = datetime.datetime.now()
    date = str(now.year) + "_" + str(now.month) + "_" + str(now.day)
    file = date + '.csv'
    wb.save(outpath+file)
    os.system("start %s" % (outpath+file))
    print ('excel_save')


#파일 MD5 해시값 얻기
def getHash_MD5(path, blocksize=65536):
    afile = open(path, 'rb')
    hasher = hashlib.md5()
    buf = afile.read(blocksize)
    while len(buf) > 0:
        hasher.update(buf)
        buf = afile.read(blocksize)
    afile.close()
    return hasher.hexdigest()

#파일 SHA1 해시값 얻기
def getHash_SHA1(path, blocksize=65536):
    afile = open(path, 'rb')
    hasher = hashlib.sha1()
    buf = afile.read(blocksize)
    while len(buf) > 0:
        hasher.update(buf)
        buf = afile.read(blocksize)
    afile.close()
    return hasher.hexdigest()